"""A portal that answers from the fixtures instead of from the network.

`TEST-03` is a Blocker: the suite never opens a browser and never reaches
SIGProv. `PortalService` takes its reader as a dependency precisely so this can
be handed in, and the whole pipeline — download, hash, `raw`, `staging`, `core`
— can be exercised with the portal switched off.
"""

import io
from pathlib import Path
from types import TracebackType
from typing import Self

from openpyxl import load_workbook

from app.modules.portal.client import (
    HTML_CONTENT_TYPE,
    PDF_CONTENT_TYPE,
    XLSX_CONTENT_TYPE,
    DownloadedDocument,
)
from app.shared.errors import ExtractionError

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures" / "portal"
PRICE_LIST = FIXTURES / "price-list-2026-08-28.xlsx"
BROKEN_LIST = FIXTURES / "price-list-broken-2026-08-28.xlsx"
HISTORY_PAGE = FIXTURES / "price-history-page-2026-08-28.html"
INVOICES_PAGE = FIXTURES / "invoices-page-2026-08-29.html"
INVOICE_FILE = FIXTURES / "invoice-F-8411-text.pdf"
SUPPLIER_LEDGER = FIXTURES / "suppliers-ledger-page-2026-08-29.html"
PURCHASE_ORDERS = FIXTURES / "purchase-orders-page-2026-08-29.html"
MESSAGES = FIXTURES / "messages-page-2026-08-31.html"
SALES = FIXTURES / "sales-page-2026-08-29.html"

PRICE_COLUMN = 5
CODE_COLUMN = 1
FIRST_DATA_ROW = 2


def price_list_bytes() -> bytes:
    """The daily file exactly as the portal published it."""
    return PRICE_LIST.read_bytes()


def broken_list_bytes() -> bytes:
    """The same file with one cell broken per row, derived by hand."""
    return BROKEN_LIST.read_bytes()


def history_page_bytes() -> bytes:
    """The history screen of `COR-0001`, as rendered."""
    return HISTORY_PAGE.read_bytes()


def invoices_page_bytes() -> bytes:
    """The invoices screen, captured from the portal with its hundred rows."""
    return INVOICES_PAGE.read_bytes()


def invoice_file_bytes() -> bytes:
    """The document of one invoice, as the supplier sent it: a PDF with a text layer."""
    return INVOICE_FILE.read_bytes()


def supplier_ledger_bytes() -> bytes:
    """The supplier register, with its eight rows already expanded."""
    return SUPPLIER_LEDGER.read_bytes()


def purchase_orders_page_bytes() -> bytes:
    """The purchase orders screen, with its forty rows."""
    return PURCHASE_ORDERS.read_bytes()


def messages_page_bytes() -> bytes:
    """The portal inbox, with its sixty-seven messages."""
    return MESSAGES.read_bytes()


def sales_page_bytes() -> bytes:
    """The sales screen, with the 588 records the survey measured."""
    return SALES.read_bytes()


# --- Rows the portal published broken -------------------------------------
#
# Four of the five screens 011 cares about were captured on a good day: only
# the sales page carries unreadable rows, twelve of them. Testing that «lo
# apartado se ve» needs a row that is actually apartada, so the broken variants
# are **derived** from the pinned pages by blanking exactly one cell — the same
# thing `price_list_with()` does to the spreadsheet, and for the same reason.
#
# Derived rather than checked in as four more `*-broken-*.html` files, because a
# second copy of a captured page drifts from the first the day the portal
# changes and nobody notices: these break a cell of whatever is pinned today, so
# they cannot describe a page that no longer exists.
#
# What makes a cell fatal is not invented either — it is read off the parser. A
# date it cannot interpret sets `reason`, and `reason` is what sends the row to
# quarantine.

UNREADABLE_CELL = "<td>—</td>"


def _blank_the_cell(page: bytes, *, before: str, cell: str, after: str = "") -> bytes:
    """Blank one cell, keeping every other cell of its row exactly as it was.

    `before` and `after` are what pin the edit to one row of one table, and they
    are also what keeps the row a row: an early version of this replaced two
    cells with one, the row came out a column short, and the payment stopped
    being a payment at all — the parser dropped it instead of quarantining it,
    which looks from the test's side exactly like the feature not working.

    The `AssertionError` is the other half of that. A silent no-op would hand
    back the intact page, no case would open, and the failure would surface far
    from here claiming the code is broken when what moved was the fixture.
    """
    html = page.decode("utf-8")
    target = before + cell + after
    assert target in html, f"the pinned page no longer has the cell {target!r} this breaks"
    return html.replace(target, before + UNREADABLE_CELL + after, 1).encode("utf-8")


def supplier_ledger_with_a_broken_payment() -> bytes:
    """The register, with one payment of Aceros Belgrano whose date is unreadable."""
    return _blank_the_cell(
        supplier_ledger_bytes(),
        before="<tr>",
        cell="<td>2026-08-15</td>",
        after="<td>Pago</td>",
    )


def supplier_ledger_with_a_broken_balance() -> bytes:
    """The register, with one supplier whose current balance is unreadable.

    The balance is the only cell of a supplier row a parser can fail on: the
    name is text and the rest of the card is optional. Until 011 the failure was
    swallowed — the reason was discarded and the row stored as readable with no
    balance — so this fixture had nothing to prove against.
    """
    return _blank_the_cell(
        supplier_ledger_bytes(),
        before="<tr><td>Aceros Belgrano SA</td>",
        cell="<td>$4.307.338</td>",
    )


def supplier_ledger_with_both_broken() -> bytes:
    """El padrón con las dos cosas rotas: el saldo de un proveedor y la fecha de un pago.

    Las dos en la **misma** lectura, porque así llegan: un documento, dos tablas
    y dos clases de pendiente. Es lo que hace falta para preguntar si las dos
    terminan en la misma pantalla (RF-06).
    """
    return _blank_the_cell(
        supplier_ledger_with_a_broken_balance(),
        before="<tr>",
        cell="<td>2026-08-15</td>",
        after="<td>Pago</td>",
    )


def purchase_orders_with_a_broken_row() -> bytes:
    """The orders screen, with one order whose date is unreadable."""
    return _blank_the_cell(
        purchase_orders_page_bytes(),
        before="<tr><td>OC-0022</td>",
        cell="<td>2026-08-07</td>",
    )


def messages_with_a_broken_row() -> bytes:
    """The inbox, with one message whose date is unreadable.

    The inbox publishes no id of its own, so a message is recognised by its date,
    sender and subject. A message whose date cannot be read therefore has no
    identity either, which is exactly why it has to be shown to somebody rather
    than merged with something else.
    """
    return _blank_the_cell(
        messages_page_bytes(),
        before="<tr>",
        cell="<td>2026-08-30</td>",
        after="<td>Insumos Industriales Bahia</td>",
    )


def price_list_with(
    *, prices: dict[str, int] | None = None, without: set[str] | None = None
) -> bytes:
    """Derive a daily file from the pinned one: change prices, drop products.

    Deriving instead of writing a spreadsheet from scratch keeps every test
    honest about the real shape of the file — six columns, a header row, prices
    as integers.
    """
    workbook = load_workbook(io.BytesIO(price_list_bytes()))
    sheet = workbook.active
    if sheet is None:  # pragma: no cover - the fixture always has a sheet
        raise ExtractionError("The fixture has no sheet")

    dropped = []
    for row in range(FIRST_DATA_ROW, sheet.max_row + 1):
        code = sheet.cell(row=row, column=CODE_COLUMN).value
        if without and code in without:
            dropped.append(row)
            continue
        if prices and code in prices:
            sheet.cell(row=row, column=PRICE_COLUMN).value = prices[code]

    for row in reversed(dropped):
        sheet.delete_rows(row)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class FakePortal:
    """A `PortalReader` that answers from bytes handed to it."""

    def __init__(
        self,
        *,
        price_list: bytes | None = None,
        history: bytes | None = None,
        invoices: bytes | None = None,
        invoice_file: bytes | None = None,
        ledger: bytes | None = None,
        fails_with: ExtractionError | None = None,
    ) -> None:
        self.price_list = price_list if price_list is not None else price_list_bytes()
        self.history = history if history is not None else history_page_bytes()
        self.invoices = invoices if invoices is not None else invoices_page_bytes()
        self.invoice_file = invoice_file if invoice_file is not None else invoice_file_bytes()
        self.supplier_ledger = ledger if ledger is not None else supplier_ledger_bytes()
        self.fails_with = fails_with
        self.downloads = 0
        self.history_visits: list[str] = []
        self.invoice_visits = 0
        self.invoice_file_visits: list[str] = []
        self.ledger_visits = 0

    def __call__(self) -> Self:
        """Usable as the `reader_factory` the service expects."""
        return self

    async def __aenter__(self) -> Self:
        return self

    async def __aexit__(
        self,
        exc_type: type[BaseException] | None,
        exc: BaseException | None,
        traceback: TracebackType | None,
    ) -> None:
        return None

    async def download_price_list(self) -> DownloadedDocument:
        """Hand back the file, or fail the way the portal fails."""
        if self.fails_with is not None:
            raise self.fails_with
        self.downloads += 1
        return DownloadedDocument(
            content=self.price_list,
            content_type=XLSX_CONTENT_TYPE,
            filename="Lista_Precios_Cordillera.xlsx",
        )

    async def fetch_product_history(self, product_code: str) -> DownloadedDocument:
        """Hand back the history screen of a product."""
        if self.fails_with is not None:
            raise self.fails_with
        self.history_visits.append(product_code)
        return DownloadedDocument(
            content=self.history,
            content_type=HTML_CONTENT_TYPE,
            filename=f"{product_code}.html",
        )

    async def fetch_invoices(self) -> DownloadedDocument:
        """Hand back the invoices screen, rendered."""
        if self.fails_with is not None:
            raise self.fails_with
        self.invoice_visits += 1
        return DownloadedDocument(
            content=self.invoices,
            content_type=HTML_CONTENT_TYPE,
            filename="facturas.html",
        )

    async def download_invoice_file(self, invoice_number: str) -> DownloadedDocument:
        """Hand back the document of one invoice.

        The visits are counted because the portal is somebody else's system with
        a shared account: a task that asks twice for a file it already has is a
        cost this side cannot see and the other side can.
        """
        if self.fails_with is not None:
            raise self.fails_with
        self.invoice_file_visits.append(invoice_number)
        return DownloadedDocument(
            content=self.invoice_file,
            content_type=PDF_CONTENT_TYPE,
            filename=f"{invoice_number}.pdf",
        )

    async def fetch_supplier_ledger(self) -> DownloadedDocument:
        """Hand back the supplier register, every row expanded."""
        if self.fails_with is not None:
            raise self.fails_with
        self.ledger_visits += 1
        return DownloadedDocument(
            content=self.supplier_ledger,
            content_type=HTML_CONTENT_TYPE,
            filename="estado-cuenta.html",
        )
