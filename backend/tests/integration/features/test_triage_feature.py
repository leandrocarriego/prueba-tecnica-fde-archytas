"""The review queue: what was set aside, and what a person decides about it.

H7 and H8 of the spec. The interesting behaviour is not that a case can be
resolved — it is that resolving it **teaches** the platform, and that undoing
that lesson gives the case back.
"""

import io
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.catalog.models import Product, ProductPrice, ProductStatus
from app.modules.identity.models import User
from app.modules.ingestion.models import ResolutionRuleProjection
from app.modules.portal.service import PortalService
from app.modules.triage.models import CaseStatus, ExceptionCase
from app.modules.triage.service import (
    MISSING_PRODUCT,
    UNKNOWN_PRODUCT,
    UNREADABLE_ROW,
    TriageService,
)
from app.shared.errors import ConflictError
from tests.factories.portal_factory import FakePortal, broken_list_bytes, price_list_with

pytestmark = [pytest.mark.integration, pytest.mark.database, pytest.mark.portal]

UNKNOWN_CODE = "COR-0999"
FIRST_PRODUCT = "COR-0001"
BROKEN_ROW_CODE = "COR-0007"


async def seed_two_runs(session: AsyncSession) -> None:
    """The first list, then one that brings a broken row and an unknown product."""
    await PortalService(session, reader_factory=FakePortal()).extract_price_list()
    await PortalService(
        session, reader_factory=FakePortal(price_list=broken_list_bytes())
    ).extract_price_list()


async def pending(session: AsyncSession, kind: str) -> list[ExceptionCase]:
    """The cases of one kind still waiting for somebody."""
    result = await session.execute(
        select(ExceptionCase).where(
            ExceptionCase.kind == kind, ExceptionCase.status == CaseStatus.PENDING
        )
    )
    return list(result.scalars().all())


async def product(session: AsyncSession, code: str) -> Product | None:
    """A product by supplier code, or None."""
    result = await session.execute(select(Product).where(Product.code == code))
    return result.scalar_one_or_none()


class TestResolvingAnUnknownProduct:
    """RF-30: it is incorporated, or left out, because a person said so."""

    async def test_incorporating_it_registers_the_product(
        self, session: AsyncSession, owner: User
    ) -> None:
        """And it shows up in the price list from then on."""
        # Arrange
        await seed_two_runs(session)
        case = (await pending(session, UNKNOWN_PRODUCT))[0]

        # Act
        await TriageService(session).resolve(
            case.id, decision={"action": "incorporate"}, user_id=owner.id
        )

        # Assert
        registered = await product(session, UNKNOWN_CODE)
        assert registered is not None
        assert registered.description == "Producto Nuevo - Articulo 999"

    async def test_the_case_leaves_the_pending_list(
        self, session: AsyncSession, owner: User
    ) -> None:
        """RF-33: the review screen has one less thing on it."""
        # Arrange
        await seed_two_runs(session)
        case = (await pending(session, UNKNOWN_PRODUCT))[0]

        # Act
        await TriageService(session).resolve(
            case.id, decision={"action": "incorporate"}, user_id=owner.id
        )

        # Assert
        assert await pending(session, UNKNOWN_PRODUCT) == []

    async def test_an_unreadable_history_can_be_given_for_reviewed(
        self, session: AsyncSession, owner: User
    ) -> None:
        """RF-45: the fourth reason a case is opened also has a way out.

        A queue with a case nobody can resolve stops emptying, and a queue that
        only grows stops being read — which takes Artículo II with it.
        """
        # Arrange
        service = TriageService(session)
        await service.open_case(
            kind="unreadable_history",
            reason="No se pudo leer el historial publicado",
            payload={"product_code": "COR-0001"},
            key="COR-0001",
        )
        await session.commit()
        case = (await pending(session, "unreadable_history"))[0]

        # Act
        resolved = await service.resolve(
            case.id, decision={"action": "ignore"}, user_id=owner.id, user_name=owner.name
        )

        # Assert
        assert resolved.status is CaseStatus.RESOLVED
        assert await pending(session, "unreadable_history") == []

    async def test_it_records_what_who_and_when(self, session: AsyncSession, owner: User) -> None:
        """RF-32: a decision without an author is not auditable.

        The name travels with the id and is stored, not looked up: the person
        who took the decision does not stop having taken it when they change
        their name or leave, and `triage` cannot read `identity` to find out.
        """
        # Arrange
        await seed_two_runs(session)
        case = (await pending(session, UNKNOWN_PRODUCT))[0]

        # Act
        resolved = await TriageService(session).resolve(
            case.id,
            decision={"action": "incorporate"},
            user_id=owner.id,
            user_name=owner.name,
        )

        # Assert
        assert resolved.status is CaseStatus.RESOLVED
        assert resolved.resolved_by_user_id == owner.id
        assert resolved.resolved_by_name == owner.name
        assert resolved.resolved_at is not None
        assert resolved.decision is not None
        assert resolved.decision["action"] == "incorporate"

    async def test_the_next_run_does_not_ask_again(
        self, session: AsyncSession, owner: User
    ) -> None:
        """RF-34: the product is known now, so there is nothing to set aside."""
        # Arrange
        await seed_two_runs(session)
        case = (await pending(session, UNKNOWN_PRODUCT))[0]
        await TriageService(session).resolve(
            case.id, decision={"action": "incorporate"}, user_id=owner.id
        )

        # Act: the same list again, with one price moved so the file is new.
        await PortalService(
            session,
            reader_factory=FakePortal(price_list=_broken_with_first_price(50_000)),
        ).extract_price_list()

        # Assert
        assert await pending(session, UNKNOWN_PRODUCT) == []

    async def test_leaving_it_out_keeps_it_out(self, session: AsyncSession, owner: User) -> None:
        """RF-30 and RF-34: the decision is "no", and it is not asked again."""
        # Arrange
        await seed_two_runs(session)
        case = (await pending(session, UNKNOWN_PRODUCT))[0]

        # Act
        await TriageService(session).resolve(
            case.id, decision={"action": "ignore"}, user_id=owner.id
        )
        await PortalService(
            session,
            reader_factory=FakePortal(price_list=_broken_with_first_price(51_000)),
        ).extract_price_list()

        # Assert
        assert await product(session, UNKNOWN_CODE) is None
        assert await pending(session, UNKNOWN_PRODUCT) == []

    async def test_the_decision_is_kept_where_ingestion_reads_it(
        self, session: AsyncSession, owner: User
    ) -> None:
        """The projection is the whole reason `ingestion` never asks `triage`."""
        # Arrange
        await seed_two_runs(session)
        case = (await pending(session, UNKNOWN_PRODUCT))[0]

        # Act
        await TriageService(session).resolve(
            case.id, decision={"action": "ignore"}, user_id=owner.id
        )

        # Assert
        rules = (await session.execute(select(ResolutionRuleProjection))).scalars().all()
        assert [rule.matcher["product_code"] for rule in rules] == [UNKNOWN_CODE]


class TestResolvingAnUnreadableRow:
    """RF-29: a person says which product and which price."""

    async def test_it_registers_the_price_that_was_indicated(
        self, session: AsyncSession, owner: User
    ) -> None:
        """The product goes back to showing a price in the list."""
        # Arrange
        await seed_two_runs(session)
        case = next(
            found
            for found in await pending(session, UNREADABLE_ROW)
            if found.payload.get("product_code") == BROKEN_ROW_CODE
        )

        # Act
        await TriageService(session).resolve(
            case.id,
            decision={"product_code": BROKEN_ROW_CODE, "price": "78914"},
            user_id=owner.id,
        )

        # Assert
        found = await product(session, BROKEN_ROW_CODE)
        assert found is not None
        price = await session.get(ProductPrice, found.id)
        assert price is not None
        assert price.price == Decimal("78914")

    async def test_the_next_run_applies_it_on_its_own(
        self, session: AsyncSession, owner: User
    ) -> None:
        """RF-34: the same broken row is not put in front of anybody twice."""
        # Arrange
        await seed_two_runs(session)
        case = next(
            found
            for found in await pending(session, UNREADABLE_ROW)
            if found.payload.get("product_code") == BROKEN_ROW_CODE
        )
        await TriageService(session).resolve(
            case.id,
            decision={"product_code": BROKEN_ROW_CODE, "price": "78914"},
            user_id=owner.id,
        )

        # Act
        await PortalService(
            session,
            reader_factory=FakePortal(price_list=_broken_with_first_price(52_000)),
        ).extract_price_list()

        # Assert
        still_pending = [
            found.payload.get("product_code") for found in await pending(session, UNREADABLE_ROW)
        ]
        assert BROKEN_ROW_CODE not in still_pending


class TestResolvingAProductThatStoppedComing:
    """RF-31: discontinued, or still in force, because a person said so."""

    async def test_it_can_be_given_up_for_discontinued(
        self, session: AsyncSession, owner: User
    ) -> None:
        """And it stops being reported run after run."""
        # Arrange
        await PortalService(session, reader_factory=FakePortal()).extract_price_list()
        await PortalService(
            session,
            reader_factory=FakePortal(price_list=price_list_with(without={FIRST_PRODUCT})),
        ).extract_price_list()
        case = (await pending(session, MISSING_PRODUCT))[0]

        # Act
        await TriageService(session).resolve(
            case.id, decision={"action": "discontinue"}, user_id=owner.id
        )

        # Assert
        found = await product(session, FIRST_PRODUCT)
        assert found is not None
        assert found.status is ProductStatus.DISCONTINUED

    async def test_it_can_be_kept_in_force_with_its_last_price(
        self, session: AsyncSession, owner: User
    ) -> None:
        """RF-31, the other half: nothing about its price is lost."""
        # Arrange
        await PortalService(session, reader_factory=FakePortal()).extract_price_list()
        await PortalService(
            session,
            reader_factory=FakePortal(price_list=price_list_with(without={FIRST_PRODUCT})),
        ).extract_price_list()
        case = (await pending(session, MISSING_PRODUCT))[0]

        # Act
        await TriageService(session).resolve(case.id, decision={"action": "keep"}, user_id=owner.id)

        # Assert
        found = await product(session, FIRST_PRODUCT)
        assert found is not None
        assert found.status is ProductStatus.ACTIVE
        price = await session.get(ProductPrice, found.id)
        assert price is not None
        assert price.price == Decimal("48210")


class TestTheRules:
    """RF-36 and RF-37: they are visible, and they can be undone."""

    async def test_a_decision_becomes_a_visible_rule(
        self, session: AsyncSession, owner: User
    ) -> None:
        """With who took it and when."""
        # Arrange
        await seed_two_runs(session)
        case = (await pending(session, UNKNOWN_PRODUCT))[0]
        await TriageService(session).resolve(
            case.id, decision={"action": "incorporate"}, user_id=owner.id
        )

        # Act
        rules = await TriageService(session).list_rules()

        # Assert
        assert len(rules) == 1
        assert rules[0].kind == UNKNOWN_PRODUCT
        assert rules[0].created_by_user_id == owner.id
        assert rules[0].revoked_at is None

    async def test_a_decision_can_be_taken_without_learning_from_it(
        self, session: AsyncSession, owner: User
    ) -> None:
        """Sometimes a case is one case, and teaching the platform would be wrong."""
        # Arrange
        await seed_two_runs(session)
        case = (await pending(session, UNKNOWN_PRODUCT))[0]

        # Act
        await TriageService(session).resolve(
            case.id, decision={"action": "ignore"}, user_id=owner.id, remember=False
        )

        # Assert
        assert await TriageService(session).list_rules() == []

    async def test_revoking_a_rule_undoes_what_it_did(
        self, session: AsyncSession, owner: User
    ) -> None:
        """RF-37: the product it incorporated stops being known."""
        # Arrange
        await seed_two_runs(session)
        case = (await pending(session, UNKNOWN_PRODUCT))[0]
        await TriageService(session).resolve(
            case.id, decision={"action": "incorporate"}, user_id=owner.id
        )
        rule = (await TriageService(session).list_rules())[0]

        # Act
        await TriageService(session).revoke_rule(rule.id, user_id=owner.id)

        # Assert
        assert await product(session, UNKNOWN_CODE) is None

    async def test_after_revoking_the_case_comes_back(
        self, session: AsyncSession, owner: User
    ) -> None:
        """RF-37: the next update sets the product aside again."""
        # Arrange
        await seed_two_runs(session)
        case = (await pending(session, UNKNOWN_PRODUCT))[0]
        await TriageService(session).resolve(
            case.id, decision={"action": "incorporate"}, user_id=owner.id
        )
        rule = (await TriageService(session).list_rules())[0]
        await TriageService(session).revoke_rule(rule.id, user_id=owner.id)

        # Act
        await PortalService(
            session,
            reader_factory=FakePortal(price_list=_broken_with_first_price(53_000)),
        ).extract_price_list()

        # Assert
        assert [
            found.payload["product_code"] for found in await pending(session, UNKNOWN_PRODUCT)
        ] == [UNKNOWN_CODE]

    async def test_a_revoked_rule_is_kept_for_the_record(
        self, session: AsyncSession, owner: User
    ) -> None:
        """It is left without effect, never deleted: a deleted rule is unauditable."""
        # Arrange
        await seed_two_runs(session)
        case = (await pending(session, UNKNOWN_PRODUCT))[0]
        await TriageService(session).resolve(
            case.id, decision={"action": "incorporate"}, user_id=owner.id
        )
        rule = (await TriageService(session).list_rules())[0]

        # Act
        await TriageService(session).revoke_rule(rule.id, user_id=owner.id)

        # Assert
        revoked = await TriageService(session).list_rules(include_revoked=True)
        assert len(revoked) == 1
        assert revoked[0].revoked_by_user_id == owner.id

    async def test_a_rule_is_not_revoked_twice(self, session: AsyncSession, owner: User) -> None:
        """The second attempt is a conflict, not a silent no-op."""
        # Arrange
        await seed_two_runs(session)
        case = (await pending(session, UNKNOWN_PRODUCT))[0]
        await TriageService(session).resolve(
            case.id, decision={"action": "incorporate"}, user_id=owner.id
        )
        rule = (await TriageService(session).list_rules())[0]
        await TriageService(session).revoke_rule(rule.id, user_id=owner.id)

        # Act / Assert
        with pytest.raises(ConflictError):
            await TriageService(session).revoke_rule(rule.id, user_id=owner.id)

    async def test_a_case_is_not_resolved_twice(self, session: AsyncSession, owner: User) -> None:
        """Two people deciding the same case is a conflict, not the last one wins."""
        # Arrange
        await seed_two_runs(session)
        case = (await pending(session, UNKNOWN_PRODUCT))[0]
        await TriageService(session).resolve(
            case.id, decision={"action": "ignore"}, user_id=owner.id
        )

        # Act / Assert
        with pytest.raises(ConflictError):
            await TriageService(session).resolve(
                case.id, decision={"action": "incorporate"}, user_id=owner.id
            )


def _broken_with_first_price(price: int) -> bytes:
    """The broken file with one known price moved, so its hash is new."""
    workbook = load_workbook(io.BytesIO(broken_list_bytes()))
    sheet = workbook.active
    assert sheet is not None
    sheet.cell(row=2, column=5).value = price
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
