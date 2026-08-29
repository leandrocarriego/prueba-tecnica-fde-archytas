"""The price update, end to end, with the portal switched off.

Five modules that do not know each other are chained by events here, and this
file drives the whole chain from its head: `PortalService` stores a document,
and by the time it returns, `ingestion` has typed it, `catalog` has applied it
and `triage` has whatever nobody could resolve. Nothing is mocked in between —
only the portal itself, and the queue.

What is asserted is the behaviour the client signed, requirement by requirement.
"""

import io
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.catalog.models import PricePoint, PriceSource, Product, ProductPrice
from app.modules.catalog.service import CatalogService
from app.modules.ingestion.models import PriceHistoryRow, PriceRow, RowStatus
from app.modules.portal.models import PortalDocument
from app.modules.portal.service import PortalService
from app.modules.triage.models import CaseStatus, ExceptionCase
from app.modules.triage.service import (
    MISSING_PRODUCT,
    UNKNOWN_PRODUCT,
    UNREADABLE_HISTORY,
    UNREADABLE_ROW,
)
from app.shared.errors import ExtractionError
from app.shared.events import NormalizedHistoryPoint
from tests.factories.portal_factory import (
    FakePortal,
    broken_list_bytes,
    price_list_with,
)
from tests.integration.features.conftest import Queued

pytestmark = [pytest.mark.integration, pytest.mark.database, pytest.mark.portal]

UNKNOWN_CODE = "COR-0999"
FIRST_PRODUCT = "COR-0001"
FIRST_PRODUCT_PRICE = Decimal("48210")


async def run_extraction(session: AsyncSession, portal: FakePortal, **kwargs: int) -> int | None:
    """Drive one extraction the way the Celery task drives it."""
    service = PortalService(session, reader_factory=portal)
    return await service.extract_price_list(**kwargs)


async def count(session: AsyncSession, model: type) -> int:
    """How many rows of something there are right now."""
    result = await session.execute(select(func.count()).select_from(model))
    return int(result.scalar_one())


async def cases_of(session: AsyncSession, kind: str) -> list[ExceptionCase]:
    """The pending cases of one kind, in the review queue."""
    result = await session.execute(
        select(ExceptionCase).where(
            ExceptionCase.kind == kind, ExceptionCase.status == CaseStatus.PENDING
        )
    )
    return list(result.scalars().all())


async def price_of(session: AsyncSession, code: str) -> ProductPrice | None:
    """The price in force of a product, by supplier code."""
    product = await session.execute(select(Product).where(Product.code == code))
    found = product.scalar_one_or_none()
    return None if found is None else await session.get(ProductPrice, found.id)


class TestTheFirstList:
    """RF-02: the first list is what establishes the catalog."""

    async def test_it_registers_every_product(self, session: AsyncSession) -> None:
        """A hundred products come in, a hundred products are known."""
        # Act
        await run_extraction(session, FakePortal())

        # Assert
        assert await count(session, Product) == 100
        assert await count(session, ProductPrice) == 100

    async def test_it_sets_aside_nothing(self, session: AsyncSession) -> None:
        """Nothing is unknown on the first list: this list is the definition."""
        # Act
        await run_extraction(session, FakePortal())

        # Assert
        assert await cases_of(session, UNKNOWN_PRODUCT) == []

    async def test_the_price_is_the_one_the_list_brought(self, session: AsyncSession) -> None:
        """RF-03, and the number the prices screen shows (RF-04)."""
        # Act
        await run_extraction(session, FakePortal())

        # Assert
        price = await price_of(session, FIRST_PRODUCT)
        assert price is not None
        assert price.price == FIRST_PRODUCT_PRICE
        assert price.is_stale is False

    async def test_the_file_is_kept_exactly_as_it_arrived(self, session: AsyncSession) -> None:
        """RF-05: the day can be explained with what the portal actually said."""
        # Arrange
        portal = FakePortal()

        # Act
        document_id = await run_extraction(session, portal)

        # Assert
        assert document_id is not None
        document = await session.get(PortalDocument, document_id)
        assert document is not None
        assert document.content == portal.price_list
        assert len(document.content_hash) == 64

    async def test_it_asks_for_the_published_history_of_each_product(
        self, session: AsyncSession, queued_history: Queued
    ) -> None:
        """RF-38, queued and spaced: a hundred visits to somebody else's system."""
        # Act
        await run_extraction(session, FakePortal())

        # Assert
        assert queued_history.count == 100
        countdowns = [call["countdown"] for call in queued_history.calls]
        assert countdowns == sorted(countdowns)
        assert countdowns[0] == 0
        assert countdowns[1] > 0


class TestTheSecondList:
    """The difference between two consecutive runs is where this breaks."""

    async def test_an_unknown_product_is_set_aside_and_not_created(
        self, session: AsyncSession
    ) -> None:
        """RF-07: the assumption may be false, and the system refuses to guess."""
        # Arrange
        await run_extraction(session, FakePortal())

        # Act
        await run_extraction(session, FakePortal(price_list=broken_list_bytes()))

        # Assert
        assert await count(session, Product) == 100
        cases = await cases_of(session, UNKNOWN_PRODUCT)
        assert [case.payload["product_code"] for case in cases] == [UNKNOWN_CODE]

    async def test_a_broken_row_does_not_stop_the_rest(self, session: AsyncSession) -> None:
        """RF-06: six rows are set aside and the other prices are registered."""
        # Arrange
        await run_extraction(session, FakePortal())

        # Act
        await run_extraction(session, FakePortal(price_list=broken_list_bytes()))

        # Assert
        assert len(await cases_of(session, UNREADABLE_ROW)) == 6
        assert await count(session, Product) == 100

    async def test_the_same_case_three_times_leaves_one_pending(
        self, session: AsyncSession
    ) -> None:
        """RF-35: the queue must not grow with the same question."""
        # Arrange
        await run_extraction(session, FakePortal())

        # Act
        for price in (61500, 61501, 61502):
            # A different price each time so the file's hash changes and the
            # run is not skipped as a duplicate.
            await run_extraction(
                session,
                FakePortal(price_list=_broken_with_price(price)),
            )

        # Assert
        cases = await cases_of(session, UNKNOWN_PRODUCT)
        assert len(cases) == 1
        assert cases[0].occurrences == 3

    async def test_a_known_product_that_stops_coming_keeps_its_price(
        self, session: AsyncSession
    ) -> None:
        """RF-08: the last price is kept, flagged, and never estimated."""
        # Arrange
        await run_extraction(session, FakePortal())

        # Act
        await run_extraction(
            session, FakePortal(price_list=price_list_with(without={FIRST_PRODUCT}))
        )

        # Assert
        price = await price_of(session, FIRST_PRODUCT)
        assert price is not None
        assert price.price == FIRST_PRODUCT_PRICE
        assert price.is_stale is True

    async def test_the_product_that_stopped_coming_is_flagged_for_review(
        self, session: AsyncSession
    ) -> None:
        """RF-28: it is put in front of a person, not given up automatically."""
        # Arrange
        await run_extraction(session, FakePortal())

        # Act
        await run_extraction(
            session, FakePortal(price_list=price_list_with(without={FIRST_PRODUCT}))
        )

        # Assert
        cases = await cases_of(session, MISSING_PRODUCT)
        assert [case.payload["product_code"] for case in cases] == [FIRST_PRODUCT]

    async def test_an_empty_list_is_a_failed_consultation(self, session: AsyncSession) -> None:
        """RF-41: a file with headers and no rows is a consultation that went wrong."""
        # Arrange
        await run_extraction(session, FakePortal())
        every_code = set((await session.execute(select(Product.code))).scalars().all())

        # Act / Assert
        with pytest.raises(ExtractionError):
            await run_extraction(
                session, FakePortal(price_list=price_list_with(without=every_code))
            )

    async def test_an_empty_list_flags_nobody_as_missing(self, session: AsyncSession) -> None:
        """RF-42: the hundred products did not stop being listed, the consultation failed."""
        # Arrange
        await run_extraction(session, FakePortal())
        every_code = set((await session.execute(select(Product.code))).scalars().all())

        # Act
        with pytest.raises(ExtractionError):
            await run_extraction(
                session, FakePortal(price_list=price_list_with(without=every_code))
            )
        await session.rollback()

        # Assert
        assert await cases_of(session, MISSING_PRODUCT) == []
        price = await price_of(session, FIRST_PRODUCT)
        assert price is not None
        assert price.is_stale is False


class TestTheEvidenceSurvivesAFailure:
    """What the portal delivered is kept even when nobody can read it.

    The day the portal changes its format is the day the parser breaks, and it
    is the same day the file is most needed to find out why. Storing the
    document and interpreting it in one transaction meant every failure of
    interpretation destroyed its own evidence (Artículo III).
    """

    async def test_a_file_that_cannot_be_read_is_still_stored(self, session: AsyncSession) -> None:
        """The empty list of RF-41 fails the run, and leaves the file behind."""
        # Arrange
        await run_extraction(session, FakePortal())
        every_code = set((await session.execute(select(Product.code))).scalars().all())

        # Act
        with pytest.raises(ExtractionError):
            await run_extraction(
                session, FakePortal(price_list=price_list_with(without=every_code))
            )
        await session.rollback()

        # Assert: the day's file is there, and it is marked as never read.
        assert await count(session, PortalDocument) == 2
        unread = (
            (
                await session.execute(
                    select(PortalDocument).where(PortalDocument.normalized_at.is_(None))
                )
            )
            .scalars()
            .all()
        )
        assert len(unread) == 1

    async def test_the_next_attempt_reads_it_instead_of_skipping_it(
        self, session: AsyncSession
    ) -> None:
        """A stored document nobody read is not a duplicate.

        Asking only "do I already have this file?" would make the retry close
        the run as successful over a file that was never interpreted — the exact
        silence RF-41 exists to break.
        """
        # Arrange: a run that fails after the file is already stored.
        await run_extraction(session, FakePortal())
        every_code = set((await session.execute(select(Product.code))).scalars().all())
        empty = FakePortal(price_list=price_list_with(without=every_code))
        with pytest.raises(ExtractionError):
            await run_extraction(session, empty)
        await session.rollback()

        # Act / Assert: the same file fails again instead of being skipped as a
        # duplicate — which is what would close the run as successful.
        with pytest.raises(ExtractionError):
            await run_extraction(session, empty)
        await session.rollback()
        assert await count(session, PortalDocument) == 2

    async def test_a_document_that_was_read_is_skipped(self, session: AsyncSession) -> None:
        """The duplicate of a file that did land is still a duplicate (`TEST-04`)."""
        # Arrange
        portal = FakePortal()
        await run_extraction(session, portal)

        # Act
        second = await run_extraction(session, portal)

        # Assert
        assert second is None
        assert await count(session, PortalDocument) == 1


class TestIdempotency:
    """`PY-07` and `TEST-04`: running it twice must change nothing twice."""

    async def test_the_same_file_is_stored_once(self, session: AsyncSession) -> None:
        """The content hash is the whole mechanism, and it is a unique index."""
        # Arrange
        portal = FakePortal()
        await run_extraction(session, portal)

        # Act
        second = await run_extraction(session, portal)

        # Assert
        assert second is None
        assert await count(session, PortalDocument) == 1
        assert await count(session, Product) == 100

    async def test_the_second_run_does_not_reprocess_anything(self, session: AsyncSession) -> None:
        """No second batch of staging rows, no duplicated points."""
        # Arrange
        portal = FakePortal()
        await run_extraction(session, portal)
        rows_after_one = await count(session, PriceRow)
        points_after_one = await count(session, PricePoint)

        # Act
        await run_extraction(session, portal)

        # Assert
        assert await count(session, PriceRow) == rows_after_one
        assert await count(session, PricePoint) == points_after_one


class TestPricesThatChange:
    """RF-22 and RF-25: what a change is, and which ones are worth looking at."""

    async def test_an_unchanged_price_adds_no_point(self, session: AsyncSession) -> None:
        """RF-22: a point is per change, not per query."""
        # Arrange
        await run_extraction(session, FakePortal())
        points_after_one = await count(session, PricePoint)

        # Act: same prices, different file (one product dropped elsewhere would
        # change the hash), so the run is processed but nothing moved.
        await run_extraction(session, FakePortal(price_list=price_list_with(prices={})))

        # Assert
        assert await count(session, PricePoint) == points_after_one

    async def test_a_changed_price_adds_a_point(self, session: AsyncSession) -> None:
        """RF-22, the other half."""
        # Arrange
        await run_extraction(session, FakePortal())

        # Act
        await run_extraction(
            session, FakePortal(price_list=price_list_with(prices={FIRST_PRODUCT: 50000}))
        )

        # Assert
        product = (
            await session.execute(select(Product).where(Product.code == FIRST_PRODUCT))
        ).scalar_one()
        points = (
            await session.execute(select(PricePoint).where(PricePoint.product_id == product.id))
        ).scalars()
        assert sorted(point.price for point in points) == [
            FIRST_PRODUCT_PRICE,
            Decimal("50000"),
        ]

    async def test_the_previous_price_is_what_the_rise_is_measured_against(
        self, session: AsyncSession
    ) -> None:
        """RF-25 compares against the previous **update**, not the previous point."""
        # Arrange
        await run_extraction(session, FakePortal())

        # Act
        await run_extraction(
            session, FakePortal(price_list=price_list_with(prices={FIRST_PRODUCT: 55441}))
        )

        # Assert
        price = await price_of(session, FIRST_PRODUCT)
        assert price is not None
        assert price.previous_price == FIRST_PRODUCT_PRICE
        # 48210 -> 55441 is +15%, above the starting 10%.
        assert price.is_highlighted is True

    async def test_a_rise_at_the_threshold_is_not_highlighted(self, session: AsyncSession) -> None:
        """ "More than 10%" is not "10%", and the edge is where this breaks."""
        # Arrange
        await run_extraction(session, FakePortal())

        # Act: exactly +10%.
        await run_extraction(
            session, FakePortal(price_list=price_list_with(prices={FIRST_PRODUCT: 53031}))
        )

        # Assert
        price = await price_of(session, FIRST_PRODUCT)
        assert price is not None
        assert price.is_highlighted is False


class TestPublishedHistory:
    """RF-38 to RF-40: what the portal already knows about a product."""

    async def test_it_lands_as_points_of_the_product(self, session: AsyncSession) -> None:
        """The eleven points the screen publishes, attached to the product."""
        # Arrange
        await run_extraction(session, FakePortal())
        portal = FakePortal()

        # Act
        await PortalService(session, reader_factory=portal).extract_product_history(FIRST_PRODUCT)

        # Assert
        product = (
            await session.execute(select(Product).where(Product.code == FIRST_PRODUCT))
        ).scalar_one()
        points = (
            await session.execute(
                select(PricePoint).where(
                    PricePoint.product_id == product.id,
                    PricePoint.source == PriceSource.PORTAL,
                )
            )
        ).scalars()
        assert len(list(points)) == 11

    async def test_importing_it_twice_leaves_the_same_points(self, session: AsyncSession) -> None:
        """RF-40, and the database is what enforces it."""
        # Arrange
        await run_extraction(session, FakePortal())
        service = CatalogService(session)
        product = (
            await session.execute(select(Product).where(Product.code == FIRST_PRODUCT))
        ).scalar_one()
        await PortalService(session, reader_factory=FakePortal()).extract_product_history(
            FIRST_PRODUCT
        )
        after_one = await _points_of(session, product.id)

        # Act: the same history again, through the service, since the second
        # extraction would be skipped by the hash before it ever gets here.
        history = await _normalized_history(session, FIRST_PRODUCT)
        await service.import_published_history(product_code=FIRST_PRODUCT, points=history)

        # Assert
        assert await _points_of(session, product.id) == after_one

    async def test_an_unreadable_history_does_not_cost_the_current_price(
        self, session: AsyncSession
    ) -> None:
        """RF-39: the point is set aside, the product keeps what it is worth."""
        # Arrange
        await run_extraction(session, FakePortal())
        unreadable = (
            b'<table class="datos"><tbody>'
            b"<tr><td>ayer</td><td>a convenir</td><td>-</td></tr>"
            b"</tbody></table>"
        )

        # Act
        await PortalService(
            session, reader_factory=FakePortal(history=unreadable)
        ).extract_product_history(FIRST_PRODUCT)

        # Assert
        price = await price_of(session, FIRST_PRODUCT)
        assert price is not None
        assert price.price == FIRST_PRODUCT_PRICE
        assert len(await cases_of(session, UNREADABLE_HISTORY)) == 1


class TestStagingIsReproducible:
    """Everything the pipeline decided is written down, not only its outcome."""

    async def test_every_line_of_the_file_leaves_a_row(self, session: AsyncSession) -> None:
        """Nothing is discarded: 101 lines in, 101 rows in `staging`."""
        # Arrange
        await run_extraction(session, FakePortal())

        # Act
        await run_extraction(session, FakePortal(price_list=broken_list_bytes()))

        # Assert: the batch number comes from a database sequence, and no
        # rollback resets a sequence, so the run is found by its document.
        latest = (await session.execute(select(func.max(PriceRow.raw_document_id)))).scalar_one()
        rows = (
            await session.execute(select(PriceRow).where(PriceRow.raw_document_id == latest))
        ).scalars()
        by_status = [row.status for row in rows]
        assert len(by_status) == 101
        assert by_status.count(RowStatus.QUARANTINED) == 6

    async def test_a_quarantined_row_keeps_what_the_file_said(self, session: AsyncSession) -> None:
        """The excerpt is what makes the review screen readable (RF-26)."""
        # Arrange
        await run_extraction(session, FakePortal())
        await run_extraction(session, FakePortal(price_list=broken_list_bytes()))

        # Act
        rows = (
            await session.execute(select(PriceRow).where(PriceRow.status == RowStatus.QUARANTINED))
        ).scalars()

        # Assert
        assert all(row.excerpt for row in rows)


# --- helpers -------------------------------------------------------------


def _broken_with_price(price: int) -> bytes:
    """The broken file with the unknown product at a different price."""

    from tests.factories.portal_factory import broken_list_bytes as broken

    workbook = load_workbook(io.BytesIO(broken()))
    sheet = workbook.active
    assert sheet is not None
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == UNKNOWN_CODE:
            sheet.cell(row=row, column=5).value = price
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def _points_of(session: AsyncSession, product_id: int) -> int:
    """How many points a product has."""
    result = await session.execute(
        select(func.count()).select_from(PricePoint).where(PricePoint.product_id == product_id)
    )
    return int(result.scalar_one())


async def _normalized_history(
    session: AsyncSession, product_code: str
) -> tuple[NormalizedHistoryPoint, ...]:
    """The points already typed in `staging`, as the event carries them."""
    rows = (
        await session.execute(
            select(PriceHistoryRow).where(
                PriceHistoryRow.product_code == product_code,
                PriceHistoryRow.status == RowStatus.VALID,
            )
        )
    ).scalars()
    return tuple(
        NormalizedHistoryPoint(staging_row_id=row.id, price=row.price, changed_at=row.changed_at)
        for row in rows
    )
