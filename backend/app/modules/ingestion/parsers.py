"""The two parsers of the price update. Pure functions, and nothing else.

They receive bytes and return typed rows. They never touch the database, never
touch the network, and never raise over a bad cell: a value that cannot be
interpreted comes back marked with the reason it could not, and quarantining it
is the service's job (`ERR-05`).

The only thing they do raise is `ExtractionError`, and only when the *document*
is not what it claims to be — a missing sheet, missing columns, no table. That
is a technical failure of the extraction, not a data problem, and it has to be
visible in `operations` instead of quietly producing zero rows.

Both are tested against the files pinned in `tests/fixtures/portal/`, never
against the live portal (`TEST-03`).
"""

import io
import re
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from html.parser import HTMLParser
from typing import Any

from openpyxl import load_workbook

from app.shared.errors import ExtractionError

# --- The daily file ------------------------------------------------------

SHEET_NAME = "Precios"
# The header the portal writes, in the order it writes it. Surveyed on
# 2026-08-29 and pinned as a fixture: `Codigo`, `Descripcion`, `Categoria`,
# `Subcategoria`, `Precio`, `Stock`.
REQUIRED_COLUMNS = ("Codigo", "Descripcion", "Precio")

# What a person reads next to the row in the review screen, so it is written in
# Spanish like everything else the user sees.
MISSING_CODE = "La fila no trae código de producto"
DUPLICATE_CODE = "El código está repetido en el archivo"
MISSING_PRICE = "La fila no trae precio"
PRICE_NOT_A_NUMBER = "El precio no es un número"
PRICE_AS_TEXT = "El precio vino como texto y no como número"
NEGATIVE_PRICE = "El precio no puede ser negativo"
UNREADABLE_DATE = "La fecha no se pudo interpretar"

DEFAULT_CURRENCY = "ARS"


@dataclass(frozen=True, slots=True)
class ParsedPriceRow:
    """One line of the daily file, as far as it could be read.

    `reason` is what says whether it could: `None` means the row is usable, and
    anything else is the sentence a person will read next to it.
    """

    line_number: int
    excerpt: str
    product_code: str | None = None
    description: str | None = None
    price: Decimal | None = None
    category_raw: str | None = None
    subcategory_raw: str | None = None
    stock: int | None = None
    reason: str | None = None

    @property
    def is_readable(self) -> bool:
        return self.reason is None


@dataclass(frozen=True, slots=True)
class ParsedHistoryPoint:
    """One row of the history screen of a product."""

    line_number: int
    excerpt: str
    price: Decimal | None = None
    changed_at: datetime | None = None
    reason: str | None = None

    @property
    def is_readable(self) -> bool:
        return self.reason is None


def _cell_text(value: object) -> str:
    """Render a cell for the excerpt, without inventing anything."""
    return "" if value is None else str(value)


def _read_price(value: object) -> tuple[Decimal | None, str | None]:
    """Interpret the price cell of the daily file.

    In this file the price is an integer — `$48.210` on screen is `48210` here,
    and the dot is a thousands separator, not a decimal point. A price that
    arrives as text is therefore **not** interpreted: `"48.210"` could be forty
    eight thousand or forty eight, and guessing between them in a system that
    controls what the business is invoiced is not a shortcut worth taking. It
    goes to a person.
    """
    if value is None or (isinstance(value, str) and not value.strip()):
        return None, MISSING_PRICE
    if isinstance(value, bool) or not isinstance(value, int | float | Decimal):
        text = str(value).strip()
        try:
            Decimal(text.replace(".", "").replace(",", "."))
        except (InvalidOperation, ArithmeticError):
            return None, PRICE_NOT_A_NUMBER
        return None, PRICE_AS_TEXT
    price = Decimal(str(value))
    if price < 0:
        return None, NEGATIVE_PRICE
    return price, None


def _read_stock(value: object) -> int | None:
    """Read the stock cell, or read nothing at all.

    A stock that cannot be read is **not** a reason to set the row aside: the
    row is about a price, and RF-06 is about the price. What cannot be read is
    simply absent, and the day that product has no photograph in the stock
    history rather than a made-up one.
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int | float | Decimal):
        return int(value)
    text = str(value).strip().replace(".", "")
    return int(text) if text.lstrip("-").isdigit() else None


def parse_price_list(content: bytes) -> list[ParsedPriceRow]:
    """Read the file of the day into rows, marking the ones that cannot be read.

    One bad cell never stops the rest: that is RF-06, and it is why this returns
    a list instead of raising.
    """
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as error:  # openpyxl raises a zoo of exceptions on bad input
        raise ExtractionError(
            "The daily file could not be opened", details={"reason": type(error).__name__}
        ) from error

    try:
        sheet = (
            workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.worksheets[0]
        )
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        raise ExtractionError("The daily file has no rows")

    header = [_cell_text(cell).strip() for cell in rows[0]]
    missing = [column for column in REQUIRED_COLUMNS if column not in header]
    if missing:
        raise ExtractionError(
            "The daily file does not have the expected columns",
            details={"missing": missing, "found": header},
        )

    index = {name: position for position, name in enumerate(header)}

    def column(row: tuple[object, ...], name: str) -> object:
        position = index.get(name)
        return None if position is None or position >= len(row) else row[position]

    parsed: list[ParsedPriceRow] = []
    seen_codes: set[str] = set()

    for offset, row in enumerate(rows[1:], start=2):
        excerpt = " | ".join(_cell_text(cell) for cell in row)
        code = _cell_text(column(row, "Codigo")).strip()
        description = _cell_text(column(row, "Descripcion")).strip()
        category = _cell_text(column(row, "Categoria")).strip() or None
        subcategory = _cell_text(column(row, "Subcategoria")).strip() or None
        stock = _read_stock(column(row, "Stock"))

        if not code and not description and not excerpt.strip(" |"):
            # A trailing empty row is not a data problem, it is the end of the
            # sheet. Nothing to set aside.
            continue

        def row_of(
            *,
            product_code: str | None = None,
            price: Decimal | None = None,
            reason: str | None = None,
            line_number: int = offset,
            excerpt: str = excerpt,
            description: str | None = description or None,
            category_raw: str | None = category,
            subcategory_raw: str | None = subcategory,
            stock: int | None = stock,
        ) -> ParsedPriceRow:
            """Build the row of this line, whatever could be read of it."""
            return ParsedPriceRow(
                line_number=line_number,
                excerpt=excerpt,
                product_code=product_code,
                description=description,
                price=price,
                category_raw=category_raw,
                subcategory_raw=subcategory_raw,
                stock=stock,
                reason=reason,
            )

        if not code:
            parsed.append(row_of(reason=MISSING_CODE))
            continue
        if code in seen_codes:
            parsed.append(row_of(product_code=code, reason=DUPLICATE_CODE))
            continue
        seen_codes.add(code)

        price, reason = _read_price(column(row, "Precio"))
        parsed.append(row_of(product_code=code, price=price, reason=reason))

    if not parsed:
        # Headers and nothing else. Letting this through would mean telling
        # `catalog` that every known product stopped being listed, which floods
        # the review queue with a hundred cases nobody caused and still closes
        # the run as successful (RF-41, RF-42). An empty list is a consultation
        # that went wrong, and it takes the same path as any other failure.
        raise ExtractionError("The daily file has no data rows")

    return parsed


# --- The published history of one product --------------------------------

# `$25.308` on the history screen: the dot separates thousands, and there are no
# cents anywhere in the hundred products the portal publishes.
MONEY = re.compile(r"^\$?\s*(-?[\d.]+)(?:,(\d{1,2}))?$")
ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


class _TableRows(HTMLParser):
    """Collects the body rows of the first `table.datos` of a page.

    A hand-written collector rather than a dependency: the portal serves one
    plain table per screen, and a parser this small does not justify pulling an
    HTML library into the build.
    """

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.rows: list[list[str]] = []
        self.found = False
        self._in_table = False
        self._done = False
        self._row: list[str] | None = None
        self._cell: list[str] | None = None
        self._is_header = False

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attributes = dict(attrs)
        if tag == "table" and not self._done:
            self._in_table = "datos" in (attributes.get("class") or "")
            self.found = self.found or self._in_table
        elif self._in_table and tag == "tr":
            self._row = []
            self._is_header = False
        elif self._in_table and tag in {"td", "th"}:
            self._cell = []
            self._is_header = self._is_header or tag == "th"

    def handle_endtag(self, tag: str) -> None:
        if not self._in_table:
            return
        if tag in {"td", "th"} and self._cell is not None and self._row is not None:
            self._row.append(" ".join("".join(self._cell).split()))
            self._cell = None
        elif tag == "tr" and self._row is not None:
            if not self._is_header:
                self.rows.append(self._row)
            self._row = None
        elif tag == "table":
            self._in_table = False
            self._done = bool(self.rows)

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)


def _read_history_price(text: str) -> tuple[Decimal | None, str | None]:
    """Interpret `$25.308` — thousands separated by dots, cents after a comma."""
    match = MONEY.match(text.strip())
    if match is None:
        return None, PRICE_NOT_A_NUMBER
    whole = match.group(1).replace(".", "")
    cents = match.group(2) or "0"
    try:
        price = Decimal(f"{whole}.{cents}")
    except (InvalidOperation, ArithmeticError):
        return None, PRICE_NOT_A_NUMBER
    if price < 0:
        return None, NEGATIVE_PRICE
    return price, None


def parse_product_history(content: bytes) -> list[ParsedHistoryPoint]:
    """Read the history screen of a product into points.

    The *Variacion vs. anterior* column the portal shows is not read: it is
    derivable from two consecutive points, and storing it would mean having two
    sources for the same number.
    """
    parser = _TableRows()
    parser.feed(content.decode("utf-8", errors="replace"))
    parser.close()

    if not parser.found:
        # No table at all: the portal changed and this parser stopped
        # understanding what it reads. That is a technical failure.
        raise ExtractionError("The history screen has no table of prices")

    if not parser.rows:
        # The table is there and publishes no price: the product has no history
        # yet. A fact, not a failure — it ends without points and without noise,
        # and the current price of the product is left alone (RF-43).
        return []

    points: list[ParsedHistoryPoint] = []
    for offset, cells in enumerate(parser.rows, start=1):
        excerpt = " | ".join(cells)
        if len(cells) < 2:
            points.append(
                ParsedHistoryPoint(line_number=offset, excerpt=excerpt, reason=UNREADABLE_DATE)
            )
            continue

        raw_date, raw_price = cells[0].strip(), cells[1].strip()
        if not ISO_DATE.match(raw_date):
            points.append(
                ParsedHistoryPoint(line_number=offset, excerpt=excerpt, reason=UNREADABLE_DATE)
            )
            continue
        try:
            changed_at = datetime.strptime(raw_date, "%Y-%m-%d").replace(tzinfo=UTC)
        except ValueError:
            points.append(
                ParsedHistoryPoint(line_number=offset, excerpt=excerpt, reason=UNREADABLE_DATE)
            )
            continue

        price, reason = _read_history_price(raw_price)
        points.append(
            ParsedHistoryPoint(
                line_number=offset,
                excerpt=excerpt,
                price=price,
                changed_at=changed_at,
                reason=reason,
            )
        )

    return points


# --- Reading a screen that has more than one table -----------------------
#
# `_TableRows` above collects the first `table.datos` of a page, which is all
# the price screens ever have. The supplier register does not fit that: it is a
# summary table, then one heading and two tables per supplier, and which
# supplier a card belongs to is said by the heading above it. So this collector
# keeps the blocks **in the order they appear**, headings included, and the
# parser walks them.


@dataclass(frozen=True, slots=True)
class Table:
    """One `table.datos` of a screen, with its header if it declared one."""

    headers: tuple[str, ...]
    rows: tuple[tuple[str, ...], ...]

    def column(self, row: tuple[str, ...], name: str) -> str:
        """The cell of this row under that header, or empty when there is none."""
        if name not in self.headers:
            return ""
        position = self.headers.index(name)
        return row[position] if position < len(row) else ""


@dataclass(frozen=True, slots=True)
class Heading:
    """A heading between two tables. In the register it names the supplier."""

    text: str


class _Blocks(HTMLParser):
    """Collects every `table.datos` and every heading of a page, in order.

    Hand-written for the same reason as `_TableRows`: the portal serves plain
    tables, and a parser this small does not justify an HTML library in the
    build. What it adds over that one is that it does not stop at the first
    table, and that it keeps the headings — the register says which supplier a
    card belongs to in an `h2`, and nowhere else.
    """

    HEADINGS = frozenset({"h1", "h2", "h3"})

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.blocks: list[Table | Heading] = []
        self.found_table = False
        self._in_table = False
        self._headers: list[str] = []
        self._rows: list[tuple[str, ...]] = []
        self._row: list[str] | None = None
        self._cell: list[str] | None = None
        self._row_is_header = False
        self._heading: list[str] | None = None
        # The register writes its supplier card as a table of `th`/`td` pairs
        # with no header row, so a `th` inside a body row is a label and not a
        # column name. Both shapes end up in `rows`, and the parser above knows
        # which it is looking at.
        self._only_th = True

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attributes = dict(attrs)
        if tag == "table":
            self._in_table = "datos" in (attributes.get("class") or "")
            self.found_table = self.found_table or self._in_table
            self._headers, self._rows = [], []
        elif tag in self.HEADINGS:
            self._heading = []
        elif self._in_table and tag == "tr":
            self._row = []
            self._row_is_header = False
            self._only_th = True
        elif self._in_table and tag in {"td", "th"}:
            self._cell = []
            self._row_is_header = self._row_is_header or tag == "th"
            self._only_th = self._only_th and tag == "th"

    def handle_endtag(self, tag: str) -> None:
        if tag in self.HEADINGS and self._heading is not None:
            self.blocks.append(Heading(" ".join("".join(self._heading).split())))
            self._heading = None
            return
        if not self._in_table:
            return
        if tag in {"td", "th"} and self._cell is not None and self._row is not None:
            self._row.append(" ".join("".join(self._cell).split()))
            self._cell = None
        elif tag == "tr" and self._row is not None:
            # A row of only `th` is the header of a table; a row that mixes
            # `th` and `td` is a labelled value, which is how the register
            # writes a supplier card.
            if self._row_is_header and self._only_th and not self._headers:
                self._headers = list(self._row)
            else:
                self._rows.append(tuple(self._row))
            self._row = None
        elif tag == "table":
            self.blocks.append(Table(headers=tuple(self._headers), rows=tuple(self._rows)))
            self._in_table = False
            self._headers, self._rows = [], []

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)
        elif self._heading is not None:
            self._heading.append(data)


def _blocks_of(content: bytes, *, section: str) -> list[Table | Heading]:
    """Read a screen into its tables and headings, or say it is not that screen."""
    parser = _Blocks()
    parser.feed(content.decode("utf-8", errors="replace"))
    parser.close()
    if not parser.found_table:
        # No table at all: the portal changed and this parser stopped
        # understanding what it reads. A technical failure of the extraction,
        # visible in `operations`, not a data problem.
        raise ExtractionError("The screen has no table", details={"section": section})
    return parser.blocks


def _tables_of(content: bytes, *, section: str) -> list[Table]:
    """The tables of a screen, without its headings."""
    return [block for block in _blocks_of(content, section=section) if isinstance(block, Table)]


# --- Reading the values a screen writes ----------------------------------

MISSING_NUMBER = "La fila no trae número"
MISSING_SUPPLIER = "La fila no trae proveedor"
MISSING_DATE = "La fila no trae fecha"
MISSING_AMOUNT = "La fila no trae monto"
AMOUNT_NOT_A_NUMBER = "El monto no es un número"
NEGATIVE_AMOUNT = "El monto no puede ser negativo"
INVALID_DATE = "La fecha no corresponde a un día que exista"
NEGATIVE_QUANTITY = "La cantidad no puede ser negativa"
MISSING_CODE_OF_SALE = "El registro no trae código de venta"


def _read_money(text: str) -> tuple[Decimal | None, str | None]:
    """Interpret `$223.376` — dots separate thousands, cents come after a comma.

    Shared by every screen that publishes an amount, because they all publish it
    the same way: measured on the invoices, the orders and the current account,
    none of which carries cents.
    """
    cleaned = text.strip()
    if not cleaned:
        return None, MISSING_AMOUNT
    match = MONEY.match(cleaned)
    if match is None:
        return None, AMOUNT_NOT_A_NUMBER
    whole = match.group(1).replace(".", "")
    cents = match.group(2) or "0"
    try:
        amount = Decimal(f"{whole}.{cents}")
    except (InvalidOperation, ArithmeticError):
        return None, AMOUNT_NOT_A_NUMBER
    if amount < 0:
        return None, NEGATIVE_AMOUNT
    return amount, None


def _read_iso_date(text: str) -> tuple[date | None, str | None]:
    """Interpret `2026-05-03`, the way every screen of the portal writes a date.

    A date that is well formed but does not exist — `2025-02-31` — is refused
    here rather than silently rolled into March: RF-17 of 009 asks for exactly
    that, and the two cases read the same to a person.
    """
    cleaned = text.strip()
    if not cleaned:
        return None, MISSING_DATE
    if not ISO_DATE.match(cleaned):
        return None, INVALID_DATE
    try:
        return datetime.strptime(cleaned, "%Y-%m-%d").date(), None
    except ValueError:
        return None, INVALID_DATE


def _read_count(text: str) -> tuple[int | None, str | None]:
    """Interpret a plain count, refusing a negative one (RF-19 of 009)."""
    cleaned = text.strip().replace(".", "")
    if not cleaned:
        return None, None
    if not cleaned.lstrip("-").isdigit():
        return None, AMOUNT_NOT_A_NUMBER
    quantity = int(cleaned)
    if quantity < 0:
        return None, NEGATIVE_QUANTITY
    return quantity, None


# `COR-0057 - Tornillos - Articulo 57`: the product cell of the invoices and the
# purchase orders screens leads with the code of the catalog, which is what lets
# a row be crossed with the catalog without guessing by name.
PRODUCT_CODE = re.compile(r"\b(COR-\d{3,})\b")


def _product_code_in(text: str) -> str | None:
    """The catalog code the cell leads with, or nothing."""
    match = PRODUCT_CODE.search(text)
    return match.group(1) if match else None


# --- The invoices screen (004) -------------------------------------------

INVOICE_COLUMNS = ("Proveedor", "Nro. Factura", "Fecha", "Monto")

# `Impaga (88d vencida)`: the state and, in red, how late it is. The overdue
# note belongs to the day it was read and not to the invoice, so only the state
# is kept — how late something is, is a question this platform answers itself.
PAYMENT_STATE = re.compile(r"^([^(]+)")

YES = frozenset({"si", "sí", "s", "true"})


@dataclass(frozen=True, slots=True)
class ParsedInvoiceRow:
    """One row of the invoices screen, as far as it could be read."""

    line_number: int
    excerpt: str
    number: str | None = None
    supplier_text: str | None = None
    issued_on: date | None = None
    due_on: date | None = None
    total: Decimal | None = None
    paid: Decimal | None = None
    balance: Decimal | None = None
    receipt_issued: bool = False
    portal_payment_status: str | None = None
    file_kind: str | None = None
    product_code: str | None = None
    reason: str | None = None

    @property
    def is_readable(self) -> bool:
        return self.reason is None


def parse_invoices(content: bytes) -> list[ParsedInvoiceRow]:
    """Read the invoices screen into rows, marking the ones that cannot be read.

    The four header data the feature needs — supplier, number, date and amount —
    are **in this table**, and the survey found not one empty cell in the
    hundred rows. That is what makes the file a second reading to compare
    against rather than the only source, and it is why a row that cannot be read
    here is a real anomaly and not the normal case.

    One bad row never stops the rest.
    """
    tables = _tables_of(content, section="invoices")
    table = tables[0]
    missing = [name for name in INVOICE_COLUMNS if name not in table.headers]
    if missing:
        raise ExtractionError(
            "The invoices screen does not have the expected columns",
            details={"missing": missing, "found": list(table.headers)},
        )

    parsed: list[ParsedInvoiceRow] = []
    for offset, row in enumerate(table.rows, start=1):
        excerpt = " | ".join(row)
        number = table.column(row, "Nro. Factura").strip()
        supplier = table.column(row, "Proveedor").strip()
        issued_on, date_reason = _read_iso_date(table.column(row, "Fecha"))
        total, amount_reason = _read_money(table.column(row, "Monto"))
        # Read but never a reason to set the row aside: an invoice with an
        # unreadable due date is still an invoice, and the due date is
        # recalculated from the supplier's agreed term anyway (RF-26 of 005).
        due_on, _ = _read_iso_date(table.column(row, "Vencimiento"))
        paid, _ = _read_money(table.column(row, "Pagado"))
        balance, _ = _read_money(table.column(row, "Saldo"))
        state = PAYMENT_STATE.match(table.column(row, "Estado de pago").strip())

        reason = (
            (None if number else MISSING_NUMBER)
            or (None if supplier else MISSING_SUPPLIER)
            or date_reason
            or amount_reason
        )
        parsed.append(
            ParsedInvoiceRow(
                line_number=offset,
                excerpt=excerpt,
                number=number or None,
                supplier_text=supplier or None,
                issued_on=issued_on,
                due_on=due_on,
                total=total,
                paid=paid,
                balance=balance,
                receipt_issued=table.column(row, "Recibo emitido").strip().lower() in YES,
                portal_payment_status=state.group(1).strip() if state else None,
                file_kind=table.column(row, "Tipo").strip() or None,
                product_code=_product_code_in(table.column(row, "Producto/insumo")),
                reason=reason,
            )
        )

    if not parsed:
        raise ExtractionError("The invoices screen has no rows")
    return parsed


# --- The supplier register (004) and its payments (005) ------------------

DETAIL_HEADING = re.compile(r"^Detalle:\s*(.+)$", re.IGNORECASE)
PAYMENT_TERM = re.compile(r"(\d+)")
MOVEMENT_COLUMNS = ("Fecha", "Tipo", "Referencia")
PAYMENT_KIND = "pago"
# The reference of a payment names its own receipt (`REC-1084`), and the
# reference of an invoice movement names the invoice (`F-8291`). Only the second
# shape can be imputed to an invoice without guessing.
INVOICE_REFERENCE = re.compile(r"\bF-\d{3,}\b")

CARD_FIELDS = {
    "cuit": "tax_id",
    "email": "email",
    "telefono": "phone",
    "teléfono": "phone",
    "condicion de pago": "payment_term_days",
    "condición de pago": "payment_term_days",
}


@dataclass(frozen=True, slots=True)
class ParsedSupplier:
    """One supplier of the register, as far as its card could be read."""

    line_number: int
    excerpt: str
    legal_name: str
    tax_id: str | None = None
    email: str | None = None
    phone: str | None = None
    payment_term_days: int | None = None
    balance: Decimal | None = None
    reason: str | None = None

    @property
    def is_readable(self) -> bool:
        return self.reason is None


@dataclass(frozen=True, slots=True)
class ParsedPayment:
    """One movement of a current account that is a payment."""

    line_number: int
    excerpt: str
    supplier_text: str
    external_id: str | None = None
    reference: str | None = None
    paid_on: date | None = None
    amount: Decimal | None = None
    reason: str | None = None

    @property
    def is_readable(self) -> bool:
        return self.reason is None


def parse_supplier_ledger(content: bytes) -> tuple[list[ParsedSupplier], list[ParsedPayment]]:
    """Read the register and the payments of every supplier that was expanded.

    The screen is a summary table and then, per supplier, a heading, a card and
    the movements of its current account. Which supplier a card belongs to is
    said by the heading above it and nowhere else, so the blocks are walked in
    order rather than picked by index.

    A supplier that was **not** expanded keeps its name and its balance and no
    card. That is a fact, not a failure: the extraction clicks every row, and if
    the portal stops opening one the platform ends up knowing less about it
    rather than inventing the rest.
    """
    blocks = _blocks_of(content, section="supplier-ledger")
    tables = [block for block in blocks if isinstance(block, Table)]
    summary = tables[0]
    if "Proveedor" not in summary.headers:
        raise ExtractionError(
            "The register does not have the expected columns",
            details={"found": list(summary.headers)},
        )

    cards: dict[str, dict[str, Any]] = {}
    payments: list[ParsedPayment] = []
    current: str | None = None
    line = 0

    for block in blocks:
        if isinstance(block, Heading):
            heading = DETAIL_HEADING.match(block.text)
            current = heading.group(1).strip() if heading else current
            continue
        if block is summary or current is None:
            continue
        if not block.headers:
            cards[current] = _read_card(block)
            continue
        if all(name in block.headers for name in MOVEMENT_COLUMNS):
            for row in block.rows:
                if block.column(row, "Tipo").strip().lower() != PAYMENT_KIND:
                    continue
                line += 1
                payments.append(_read_movement(block, row, supplier=current, line_number=line))

    suppliers: list[ParsedSupplier] = []
    for offset, row in enumerate(summary.rows, start=1):
        name = summary.column(row, "Proveedor").strip()
        if not name:
            continue
        # El motivo **viaja**. Hasta la 011 esta línea decía `balance, _`: un
        # saldo que no se podía interpretar se guardaba como `None` y la fila
        # quedaba marcada legible, así que nunca iba a cuarentena, nunca abría
        # caso y nadie se enteraba de que el padrón había publicado un número
        # ilegible. Es exactamente lo que el Artículo II prohíbe —descartar en
        # silencio— un nivel más abajo de donde la 011 lo vino a cerrar: la
        # feature agregó el evento y el suscriptor, y el evento no podía
        # dispararse porque acá se tiraba el único dato que lo dispara (RF-01).
        balance, balance_reason = _read_money(summary.column(row, "Saldo actual"))
        card = cards.get(name, {})
        suppliers.append(
            ParsedSupplier(
                line_number=offset,
                excerpt=" | ".join(row),
                legal_name=name,
                tax_id=card.get("tax_id"),
                email=card.get("email"),
                phone=card.get("phone"),
                payment_term_days=card.get("payment_term_days"),
                balance=balance,
                reason=balance_reason,
            )
        )

    if not suppliers:
        raise ExtractionError("The register has no suppliers")
    return suppliers, payments


def _read_card(table: Table) -> dict[str, Any]:
    """Read the labelled card of one supplier: tax id, email, phone, term.

    Looked up by label and not by position: the card is the one place the tax id
    exists, and a column that moves must not silently turn a phone into a term.
    """
    card: dict[str, Any] = {}
    for row in table.rows:
        if len(row) < 2:
            continue
        field = CARD_FIELDS.get(row[0].strip().lower())
        if field is None:
            continue
        value = row[1].strip()
        if field == "payment_term_days":
            days = PAYMENT_TERM.search(value)
            card[field] = int(days.group(1)) if days else None
        else:
            card[field] = value or None
    return card


def _read_movement(
    table: Table, row: tuple[str, ...], *, supplier: str, line_number: int
) -> ParsedPayment:
    """Read one payment of a current account, without deciding whose invoice it is.

    The reference is kept exactly as the portal wrote it. Whether it names an
    invoice is not a question for a parser: a voucher that names one is imputed,
    and one that names its own receipt number waits for a person to say which
    invoices it covers (RF-53 of 005). Guessing here would be the platform
    splitting money on its own, which is the one thing H2 forbids.
    """
    excerpt = " | ".join(row)
    reference = table.column(row, "Referencia").strip()
    paid_on, date_reason = _read_iso_date(table.column(row, "Fecha"))
    amount, amount_reason = _read_money(table.column(row, "Haber"))
    return ParsedPayment(
        line_number=line_number,
        excerpt=excerpt,
        supplier_text=supplier,
        # The receipt number of the portal is what makes the same voucher the
        # same voucher between two runs, so a second reading imputes it once.
        external_id=f"{supplier}|{reference}" if reference else None,
        reference=reference or None,
        paid_on=paid_on,
        amount=amount,
        reason=date_reason or amount_reason,
    )


def invoice_references_in(reference: str) -> list[str]:
    """The invoice numbers a voucher's reference names, if it names any."""
    return INVOICE_REFERENCE.findall(reference or "")


# --- The purchase orders screen (007) ------------------------------------

ORDER_COLUMNS = ("Nro. OC", "Fecha", "Proveedor", "Estado")


@dataclass(frozen=True, slots=True)
class ParsedPurchaseOrder:
    """One row of the purchase orders screen, as far as it could be read."""

    line_number: int
    excerpt: str
    number: str | None = None
    ordered_on: date | None = None
    supplier_text: str | None = None
    product_code: str | None = None
    product_text: str | None = None
    quantity: int | None = None
    amount: Decimal | None = None
    status_text: str | None = None
    reason: str | None = None

    @property
    def is_readable(self) -> bool:
        return self.reason is None


def parse_purchase_orders(content: bytes) -> list[ParsedPurchaseOrder]:
    """Read the purchase orders screen into rows.

    The product cell leads with the code of the catalog and links to its page,
    so an order crosses with the catalog without guessing by name. The screen
    publishes **one date**, the order's: since when an order has been in its
    state is not there, and the platform only knows it from the moment it starts
    observing it itself (RF-05, RF-48 of 007).
    """
    table = _tables_of(content, section="purchase-orders")[0]
    missing = [name for name in ORDER_COLUMNS if name not in table.headers]
    if missing:
        raise ExtractionError(
            "The purchase orders screen does not have the expected columns",
            details={"missing": missing, "found": list(table.headers)},
        )

    parsed: list[ParsedPurchaseOrder] = []
    for offset, row in enumerate(table.rows, start=1):
        number = table.column(row, "Nro. OC").strip()
        supplier = table.column(row, "Proveedor").strip()
        ordered_on, date_reason = _read_iso_date(table.column(row, "Fecha"))
        amount, _ = _read_money(table.column(row, "Monto estimado"))
        quantity, _ = _read_count(table.column(row, "Cantidad"))
        product = table.column(row, "Producto/insumo").strip()
        status = table.column(row, "Estado").strip()

        parsed.append(
            ParsedPurchaseOrder(
                line_number=offset,
                excerpt=" | ".join(row),
                number=number or None,
                ordered_on=ordered_on,
                supplier_text=supplier or None,
                product_code=_product_code_in(product),
                product_text=product or None,
                quantity=quantity,
                amount=amount,
                status_text=status or None,
                reason=(None if number else MISSING_NUMBER)
                or (None if supplier else MISSING_SUPPLIER)
                or date_reason
                or (None if status else "La fila no trae estado"),
            )
        )

    if not parsed:
        raise ExtractionError("The purchase orders screen has no rows")
    return parsed


# --- The inbox of the portal (007) ---------------------------------------

MESSAGE_COLUMNS = ("Fecha", "Remitente", "Asunto")
READ_STATES = frozenset({"leido", "leído"})
# What separates the kind from the rest of a subject: «Reclamo de pago - F-1809».
SUBJECT_KIND_SEPARATOR = " - "


@dataclass(frozen=True, slots=True)
class ParsedMessage:
    """One message of the inbox, as far as it could be read."""

    line_number: int
    excerpt: str
    external_id: str | None = None
    received_at: datetime | None = None
    sender_text: str | None = None
    kind_text: str | None = None
    subject: str | None = None
    body: str | None = None
    already_read: bool = False
    reason: str | None = None

    @property
    def is_readable(self) -> bool:
        return self.reason is None


def _kind_written_in(column: str, subject: str) -> str | None:
    """How the portal names the kind of a message, wherever it writes it.

    **There is no `Tipo` column**, and this used to read one: verified against
    the live inbox on 2026-08-31, whose columns are `Fecha`, `Remitente`,
    `Asunto` and `Estado`. Every message came back with no kind at all, so all
    sixty-seven were shown unclassified and **no immediate alert could ever
    fire** — RF-33 and RF-34 were dead without anything failing.

    The kind is in the subject, before a dash: «Reclamo de pago - F-1809»,
    «Vencimiento proximo - F-4032», «Stock bajo - COR-0143». The column is still
    read first, because a portal that grows one should be believed over a
    convention, and this stays a *reading* rather than a decision: what it means
    is `messaging`'s to say, and a wording nobody mapped is still shown
    unclassified (RF-25).
    """
    named = column.strip()
    if named:
        return named
    head = subject.split(SUBJECT_KIND_SEPARATOR)[0].strip()
    return head or None


def parse_messages(content: bytes) -> list[ParsedMessage]:
    """Read the inbox into messages.

    A message whose kind cannot be determined is **not** set aside: it is shown
    unclassified, which is what RF-25 asks for. What does set one aside is not
    being able to tell it apart from another — no id and no date — because two
    readings of the inbox would then register it twice.
    """
    table = _tables_of(content, section="messages")[0]
    missing = [name for name in MESSAGE_COLUMNS if name not in table.headers]
    if missing:
        raise ExtractionError(
            "The inbox does not have the expected columns",
            details={"missing": missing, "found": list(table.headers)},
        )

    parsed: list[ParsedMessage] = []
    for offset, row in enumerate(table.rows, start=1):
        excerpt = " | ".join(row)
        identifier = table.column(row, "Id").strip() or table.column(row, "Nro.").strip()
        received, date_reason = _read_iso_date(table.column(row, "Fecha"))
        sender = table.column(row, "Remitente").strip()
        subject = table.column(row, "Asunto").strip()
        state = table.column(row, "Estado").strip().lower()

        parsed.append(
            ParsedMessage(
                line_number=offset,
                excerpt=excerpt,
                # Without an id of its own, the message is identified by what it
                # says and when: enough for the same message to be recognised
                # between two readings, and never enough to merge two different
                # ones — the date, the sender and the subject would all have to
                # coincide.
                external_id=identifier or f"{received}|{sender}|{subject}",
                received_at=(
                    None
                    if received is None
                    else datetime.combine(received, datetime.min.time(), tzinfo=UTC)
                ),
                sender_text=sender or None,
                kind_text=_kind_written_in(table.column(row, "Tipo"), subject),
                subject=subject or None,
                body=table.column(row, "Mensaje").strip() or subject or None,
                already_read=state in READ_STATES,
                reason=date_reason if not identifier else None,
            )
        )

    if not parsed:
        raise ExtractionError("The inbox has no rows")
    return parsed


# --- The sales screen (009) ----------------------------------------------

SALE_COLUMNS = ("Codigo", "Fecha", "Total")

# Two records are the same sale when their codes differ only in spelling:
# spaces, dashes and case. Nothing else is collapsed — a code that differs in a
# digit is a different sale, and no normalisation may hide that.
SALE_CODE_NOISE = re.compile(r"[\s\-_.]+")


def sale_code_key(code: str) -> str:
    """The sale code with the differences of spelling removed (RF-10 of 009)."""
    return SALE_CODE_NOISE.sub("", code).casefold()


@dataclass(frozen=True, slots=True)
class ParsedSale:
    """One sales record, as far as it could be read."""

    line_number: int
    excerpt: str
    code: str | None = None
    code_key: str | None = None
    sold_on: date | None = None
    product_code: str | None = None
    quantity: int | None = None
    total: Decimal | None = None
    reason: str | None = None

    @property
    def is_readable(self) -> bool:
        return self.reason is None


def parse_sales(content: bytes) -> list[ParsedSale]:
    """Read the sales screen into records, setting aside the ones no total may add.

    Everything RF-16 to RF-19 of 009 names is decided here, and nothing is ever
    completed by assumption (RF-24): a missing date, a date that does not exist,
    a missing total and a negative quantity each come back with their own
    reason, and the record waits for a person instead of being added up.
    """
    table = _tables_of(content, section="sales")[0]
    missing = [name for name in SALE_COLUMNS if name not in table.headers]
    if missing:
        raise ExtractionError(
            "The sales screen does not have the expected columns",
            details={"missing": missing, "found": list(table.headers)},
        )

    parsed: list[ParsedSale] = []
    for offset, row in enumerate(table.rows, start=1):
        code = table.column(row, "Codigo").strip()
        sold_on, date_reason = _read_iso_date(table.column(row, "Fecha"))
        total, amount_reason = _read_money(table.column(row, "Total"))
        quantity, quantity_reason = _read_count(table.column(row, "Cantidad"))
        product = table.column(row, "Producto").strip()

        parsed.append(
            ParsedSale(
                line_number=offset,
                excerpt=" | ".join(row),
                code=code or None,
                code_key=sale_code_key(code) if code else None,
                sold_on=sold_on,
                product_code=_product_code_in(product),
                quantity=quantity,
                total=total,
                reason=(None if code else MISSING_CODE_OF_SALE)
                or date_reason
                or amount_reason
                or quantity_reason,
            )
        )

    if not parsed:
        raise ExtractionError("The sales screen has no rows")
    return parsed
