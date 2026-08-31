"""Reading the document of an invoice: a PDF, a scan, or a spreadsheet.

Pure functions over bytes, like `parsers.py`, and with the same rule: nothing
here raises over a value it cannot read. What could not be read comes back
marked, and setting the invoice aside is the service's job (`ERR-05`).

**Why three readers and no model.** The survey measured the three formats the
portal publishes — 25 PDFs with a text layer, 46 scans, 29 spreadsheets — and
the client refused a per-document cost (brief 1.4.0), which rules out a paid OCR
service and an LLM per invoice. `research.md` has the measurement. So: `pypdf`
for the text layer, Tesseract for the scan, `openpyxl` for the spreadsheet.

**What the reading is for.** Not to be the source of the data — the invoices
table already publishes the four header fields, complete, in all hundred rows.
It is the *second* reading, and what matters is whether the two agree: when they
do the invoice is certainty, and when they do not it goes to a person with the
excerpt in view. That comparison replaces the confidence threshold that would
otherwise be the only thing separating a good number from a made-up one.

**The trap.** The only tax id printed on these documents is **Cordillera's**,
the client's, never the issuer's. A reader that keeps the first one it finds
assigns the same supplier to all hundred invoices. So no tax id is read here at
all: the supplier travels as the name the document prints, and resolving it is
`purchases`' problem, against the register.
"""

import io
import re
from collections.abc import Callable
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader

from app.logging import get_logger

logger = get_logger(__name__)

# Inside the document the date is `03/05/2026` — day first — and **not** ISO
# like the table. Measured on the four pinned invoices; getting this backwards
# turns the third of May into the fifth of March and every comparison disagrees.
DOCUMENT_DATE = re.compile(r"\b(\d{1,2})/(\d{1,2})/(\d{4})\b")
INVOICE_NUMBER = re.compile(r"\b(F-\d{3,})\b")
AMOUNT = re.compile(r"\$\s*([\d.]+)(?:,(\d{1,2}))?")
# A spreadsheet invoice writes its amounts as numbers, with no `$` anywhere, so
# the total is found by its label instead. Both shapes are looked for, and the
# labelled one wins when it is there: it is the document saying which number is
# the total, rather than this reader deciding.
LABELLED_TOTAL = re.compile(
    r"(?:monto\s+total|total)\s*[:\-]?\s*\$?\s*([\d.]+)(?:,(\d{1,2}))?", re.IGNORECASE
)
SUPPLIER_LABEL = re.compile(r"proveedor\s*[:\-]\s*(.+)", re.IGNORECASE)

# The languages the OCR is asked for, in order. Spanish is what the documents
# are written in and what the worker image installs; the fallback is there so a
# machine without the language pack reads *something* — a reading that then
# disagrees with the table and sends the invoice to a person, which is the safe
# end of the mistake.
OCR_LANGUAGES = ("spa", "eng")

# One reader: bytes in, whatever text it could get out. Named so the order they
# are tried in is a list of them and not a list of `object`.
type Reader = Callable[[bytes], str]

UNREADABLE_DOCUMENT = "No se pudo leer el archivo de la factura"
NOTHING_IN_THE_DOCUMENT = "El archivo no dice número, fecha ni monto"
EXCERPT_LIMIT = 2000


@dataclass(frozen=True, slots=True)
class DocumentReading:
    """What the document itself said, as far as it could be read."""

    readable: bool
    excerpt: str
    number: str | None = None
    issued_on: date | None = None
    total: Decimal | None = None
    supplier_text: str | None = None
    reason: str | None = None

    def agrees_with(
        self, *, number: str | None, issued_on: date | None, total: Decimal | None
    ) -> bool:
        """Whether this reading says the same as the table about the same invoice.

        A field the document does not carry is not a disagreement: it is one
        fewer confirmation. What counts as a disagreement is the document
        stating something **different** — and that is exactly what sends the
        invoice to a person (RF-29, RF-30 of 004).
        """
        if not self.readable:
            return False
        for read, expected in (
            (self.number, number),
            (self.issued_on, issued_on),
            (self.total, total),
        ):
            if read is not None and expected is not None and read != expected:
                return False
        # A document that confirmed nothing at all is not agreement either.
        return any(value is not None for value in (self.number, self.issued_on, self.total))


def read_invoice_document(
    content: bytes, *, content_type: str = "", file_kind: str = ""
) -> DocumentReading:
    """Read an invoice document with whichever reader understands it.

    The kind the table declares picks who tries first, and the others try
    afterwards: the column is a hint, not a promise, and a scan mislabelled as a
    PDF must not cost the invoice its reading.
    """
    for reader in _readers_for(content_type=content_type, file_kind=file_kind):
        try:
            text = reader(content)
        except Exception as error:  # every reader has its own zoo of failures
            logger.warning(
                "An invoice reader failed",
                extra={"reader": reader.__name__, "error": type(error).__name__},
            )
            continue
        if text and text.strip():
            return _fields_in(text)

    return DocumentReading(readable=False, excerpt="", reason=UNREADABLE_DOCUMENT)


def _readers_for(*, content_type: str, file_kind: str) -> list[Reader]:
    """The readers to try, best guess first."""
    kind = f"{file_kind} {content_type}".lower()
    if "excel" in kind or "spreadsheet" in kind or "xlsx" in kind:
        return [_from_spreadsheet, _from_pdf_text, _from_scan]
    if "escane" in kind or "scan" in kind:
        return [_from_scan, _from_pdf_text, _from_spreadsheet]
    return [_from_pdf_text, _from_scan, _from_spreadsheet]


def _from_pdf_text(content: bytes) -> str:
    """The text layer of a PDF, when it has one."""
    reader = PdfReader(io.BytesIO(content))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _from_scan(content: bytes) -> str:
    """The text of a scanned PDF, read from the images it embeds.

    The images are pulled out with `pypdf` instead of rendering the page. That
    avoids a second system dependency for the rendering, and it is exactly right
    for what the portal publishes: a scan here is one JPEG per page, and the
    embedded image *is* the page.
    """
    import pytesseract
    from PIL import Image

    reader = PdfReader(io.BytesIO(content))
    pieces: list[str] = []
    for page in reader.pages:
        for embedded in page.images:
            image = Image.open(io.BytesIO(embedded.data))
            pieces.append(_ocr(image, pytesseract))
    return "\n".join(pieces)


def _ocr(image: Any, pytesseract: Any) -> str:
    """Run the OCR, falling back to another language rather than to nothing."""
    last: Exception | None = None
    for language in OCR_LANGUAGES:
        try:
            return str(pytesseract.image_to_string(image, lang=language))
        except Exception as error:  # a missing language pack, mostly
            last = error
    logger.warning("The OCR could not read a page", extra={"error": type(last).__name__})
    return ""


def _from_spreadsheet(content: bytes) -> str:
    """Every cell of a spreadsheet invoice, as lines of `label: value`.

    The header of these files sits at `A2`, with blank rows in between and a
    `TOTAL` line at the foot, so nothing is read by position: the whole sheet is
    flattened and the labels are looked up by name in the text, exactly as they
    are in a PDF. One reader for three shapes instead of three.
    """
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        lines: list[str] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell) for cell in row if cell is not None]
                if cells:
                    lines.append(": ".join(cells) if len(cells) > 1 else cells[0])
    finally:
        workbook.close()
    return "\n".join(lines)


def _fields_in(text: str) -> DocumentReading:
    """Pull the four header fields out of whatever the readers produced."""
    excerpt = text.strip()[:EXCERPT_LIMIT]
    number = INVOICE_NUMBER.search(text)
    issued_on = _date_in(text)
    total = _total_in(text)
    supplier = SUPPLIER_LABEL.search(text)

    reading = DocumentReading(
        readable=True,
        excerpt=excerpt,
        number=number.group(1) if number else None,
        issued_on=issued_on,
        total=total,
        supplier_text=supplier.group(1).strip() if supplier else None,
    )
    if reading.number is None and reading.issued_on is None and reading.total is None:
        # The bytes were read and say nothing this platform recognises. That is
        # not a crash and not a silent success: it is an invoice whose document
        # confirms nothing, and it goes to a person.
        return DocumentReading(readable=True, excerpt=excerpt, reason=NOTHING_IN_THE_DOCUMENT)
    return reading


def _date_in(text: str) -> date | None:
    """The first `dd/mm/aaaa` of the document, read day first."""
    match = DOCUMENT_DATE.search(text)
    if match is None:
        return None
    day, month, year = (int(part) for part in match.groups())
    try:
        return datetime(year, month, day).date()
    except ValueError:
        return None


def _total_in(text: str) -> Decimal | None:
    """The largest amount the document prints.

    A total the document **labels** wins: that is the document saying which
    number is the total, rather than this reader deciding. Failing that, the
    largest printed amount — a spreadsheet prints the line and then the sum, a
    PDF prints the total alone, and the biggest is right in both shapes. Where
    it is not, the number disagrees with the table and the invoice goes to a
    person instead of being believed.
    """
    labelled = [_as_amount(whole, cents) for whole, cents in LABELLED_TOTAL.findall(text)]
    named = [amount for amount in labelled if amount is not None]
    if named:
        return max(named)
    amounts = [_as_amount(whole, cents) for whole, cents in AMOUNT.findall(text)]
    printed = [amount for amount in amounts if amount is not None]
    return max(printed) if printed else None


def _as_amount(whole: str, cents: str) -> Decimal | None:
    """One printed amount, or nothing when it is not a number after all."""
    try:
        return Decimal(f"{whole.replace('.', '').strip('.') or 0}.{cents or 0}")
    except (InvalidOperation, ArithmeticError):
        return None
